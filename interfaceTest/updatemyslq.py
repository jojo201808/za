import pymysql
import openpyxl
import os

def connectMysql(a):
    db = pymysql.connect(host='106.14.193.137',user='bubble',password='TestBubble789#!',db='bubble',port=6603)
    cursor = db.cursor()    
    b = []
    for i in a:
        sql = "update mp_user_wallet where id='%s'" % i
        cursor.execute(sql)
        record = cursor.fetchone()
        b.append(record[0])
    db.close()
    return b



def readExcel():
    path = os.path.abspath('.') + '\\testFile\\123.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    rows = sheet.max_row
    a = []
    for i in range(1,rows+1):
        svalue = sheet.cell(i,1).value
        if svalue != None:
            a.append(svalue)
    connectMysql(a)
    wb.save(path)

if __name__=='__main__':
    readExcel()