#代理后台
import requests,os
#import sys
#BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#sys.path.append(BASE_DIR)
import setingproxy
import json
import openpyxl
import time
import random
def parseJson(s):
    try:
        d = json.loads(s)
    except Exception as e:
        print(str(e))
        return None
    return d

def getToken():
    r = requests.post(
        url='https://proxy-manager-test.maopp.cn/user/login',
        data='{"username":"15667447563","password":"aaaaaa"}',
        headers=setingproxy.HEADERS)
    t = parseJson(r.text)
    return t['data']['token']

def test1():
    path = os.path.abspath('.') + '\\testFile\\testdataproxy.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    rows = sheet.max_row
    token1 = getToken()
    for i in range(2,rows+1):
        time.sleep(1+random.random())
        tmp = sheet.cell(i,1).value
        if tmp != None:           
            ifurl = sheet.cell(i,4).value
            #userid = sheet.cell(i,6).value
            ip = sheet.cell(i,3).value
            #version = sheet.cell(i,7).value
            data = sheet.cell(i,5).value
            tmp = parseJson(str(data))
            if tmp == None:
                print("第%s行，第5列json格式不对" % i)
                continue
            
            expectresult = parseJson(str(sheet.cell(i,8).value))
            if expectresult == None:
                print("第%s行，第8列json格式不对" % i)
                continue

            url = ip + ifurl
            headers = setingproxy.HEADERS
            #headers['userId'] = userid
            #headers['appVersion'] = version
            headers['token'] = token1

            r = requests.post(url,data=data.encode('utf-8'),headers=headers)
            rdic = parseJson(r.text)
            if rdic == None:
                sheet.cell(i,10).value = 'FAIL'
                sheet.cell(i,11).value = r.text
                continue
            d,n = compare(expectresult,rdic)

            sheet.cell(i,9).value = json.dumps(rdic,ensure_ascii=False,indent=4)
            if n==0:
                sheet.cell(i,10).value = 'PASS'
            else:
                sheet.cell(i,10).value = 'FAIL'
                sheet.cell(i,11).value = json.dumps(d,ensure_ascii=False,indent=4)
    wb.save(path)

def compare(d1,d2):
    d = {}
    n = 0
    def compareSimple(dict1, dict2):
        for k,v in dict1.items():
            if dict2.get(k) == None or dict2.get(k) != v:
                return False
        return True

    def compareV(dict1,dict2,result):
        nonlocal d,n
        for k,v in dict1.items():
            if isinstance(v,dict):
                if isinstance(dict2[k],dict):    
                    if result.get(k) == None:
                        result[k] = {}
                        compareV(dict1[k],dict2[k],result[k])
                else:
                    result[k] = dict2[k]
                    n = n + 1

            elif isinstance(v,list):
                if isinstance(dict2[k],list):
                    if result.get(k) ==None:
                        result[k] = []
                    tmp = False
                    for i in range(len(v)):
                        tmp = False
                        for j in range(len(dict2[k])):
                            tmp = compareSimple(dict1[k][i],dict2[k][j])
                            if tmp == True:
                                break
                        if tmp == False:
                            n = n + 1
                            result[k]=dict2[k]                       

            else:
                if dict1[k] != dict2[k]:
                    result[k] = dict2[k]
                    n = n + 1
    compareV(d1, d2, d)
    return d,n

if __name__ == '__main__':
    test1()