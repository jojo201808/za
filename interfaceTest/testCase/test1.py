import requests,os
import sys
#BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#sys.path.append(BASE_DIR)
import settings1
import json
import openpyxl

def test1():
    path = os.path.abspath('.') + '\\testFile\\testdata.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    rows = sheet.max_row
    for i in range(2,rows+1):
        data = sheet.cell(i,5).value
        ifurl = sheet.cell(i,4).value
        userid = sheet.cell(i,6).value
        ip = sheet.cell(i,3).value
        expectresult = json.loads(str(sheet.cell(i,7).value))

        url = ip + ifurl
        headers = settings1.HEADERS
        headers['userId'] = userid
        
        r = requests.post(url,data=data,headers=headers)
        rdic = r.json()
        d,n = compare(expectresult,rdic)
        

        sheet.cell(i,8).value = json.dumps(rdic,ensure_ascii=False,indent=4)
        if n==0:
            sheet.cell(i,9).value = 'PASS'
        else:
            sheet.cell(i,9).value = 'FAIL'
            sheet.cell(i,10).value = json.dumps(d,ensure_ascii=False,indent=4)
    wb.save(path)

def compare(d1,d2):
    d = {}
    n = 0
    def compareV(dict1,dict2,result):
        nonlocal d,n
        for k,v in dict1.items():
            if isinstance(v,dict):
                if isinstance(dict2[k],dict):    
                    if result.get(k) == None:
                        result[k] = {}
                        compareV(dict1[k],dict2[k],result[k])
                else:
                    result[k] = dict2[k]
                    n = n + 1
            else:
                if dict1[k] != dict2[k]:
                    result[k] = dict2[k]
                    n = n + 1

    compareV(d1, d2, d)
    return d,n

if __name__ == '__main__':
    test1()